import csv
import datetime as dt
import hashlib
import json
import os
import shutil
import openpyxl as openpyxl
import codecs
from collections import defaultdict
from typing import Union, List, Any, Dict
import chardet

"""
This class is a storage locker for function I use to convert files, check file info
    
"""


class Utility:
    # noinspection PyUnboundLocalVariable
    @classmethod
    def naming(
        cls, descriptor: str, lang: str, length: int, ext: str, long_date: bool = True
    ) -> str:
        """
        function to name files given a few features
        :param descriptor: short descriptor of the file use-case like `black_cat_hats`
        :param lang: 4 char language code like `en_us`
        :param length: length of output, if not given will not show
        :param ext: extension to go with name
        :param long_date: bool to denote if filename should include just date, or date time, defaults to just date
        :return new_name: name converted from input
        """
        # Naming convention based on vars given

        if lang is None:
            new_name = f"{descriptor}-{length}-{cls.now(long_date)}.{ext}"
        elif lang is None and length is None:
            new_name = f"{descriptor}-{cls.now(long_date)}.{ext}"
        elif descriptor is None:
            new_name = f"Data-{lang}-{length}-{cls.now(long_date)}.{ext}"
        elif descriptor is None and lang is None:
            new_name = f"Data-{length}-{cls.now(long_date)}.{ext}"
        elif descriptor is None and length is None:
            new_name = f"Data-{lang}-{cls.now(long_date)}.{ext}"
        elif length is None:
            new_name = f"{descriptor}-{lang}-{cls.now(long_date)}.{ext}"
        else:
            new_name = f"{descriptor}-{lang}-{length}-{cls.now(long_date)}.{ext}"

        return new_name

    @staticmethod
    def group_data(
        in_array: list, var: str, group_size: int = 3
    ) -> tuple[Union[dict[list], dict], Union[dict[list], dict]]:
        """
        Function to group unites based on dict key variable
        and return all, and those in groupings of specified group size
        :param in_array: an array of dict objects
        :param group_size: minimum size needed to return groupings
        :param var: variable to use as keyword to be searched in dict keys
        :return: out - group with size info, groups - all items
        """
        groups = defaultdict(list)
        out = defaultdict(list)

        for i in in_array:
            groups[i[var]].append(i)

        for k, v in dict(groups).items():
            if len(v) >= group_size:
                out[k] = v
                continue

        print(f"Total unique items: {len(in_array)}")
        print(f"Total unique groupings: {len(groups)}")
        print(f"Total unique groupings of {group_size}  or more: {len(out)}")
        return dict(out), dict(groups)

    @classmethod
    def cycle_through(
        cls,
        location: Union[str, bytes, os.PathLike] = "./",
        new_location: Union[str, bytes, os.PathLike] = "./",
        known_unique: dict = None,
        ignore: dict = None,
    ) -> Dict:
        """
        Script to copy data from a multitude of folders to new single location, with only unique filenames
        :param location: Location of the input data to run the script over
        :param new_location: Location to copy the data to
        :param known_unique: complete dict of already known items
        :param ignore: array of files to ignore
        :return: updated unique_id list
        """
        known_unique = known_unique if known_unique else {}
        ignore = ignore if ignore else {}

        for root, dirs, files in os.walk(location):
            for c, i in enumerate(files):
                if i == "" or i in ignore:
                    continue
                # create dict for item if it doesn't already exist
                if i not in list(known_unique.keys()):
                    known_unique[i] = {"old_path": root}
                # Check if the item associated with the user is in the collection already
                try:
                    if i in list(known_unique[i].keys()):
                        continue
                except TypeError:
                    pass
                pathing = root.split("\\")[-3:]
                pathing_name = "-".join(pathing[1:])
                new_name = f"{pathing_name}-{i}"
                cls.copy_organize(
                    file_name=i,
                    old_path=root,
                    new_path=new_location,
                    new_filename=new_name,
                )
                known_unique[i]["new_path"] = new_location
                known_unique[i]["new_name"] = new_name

        return known_unique

    @staticmethod
    def copy_organize(
        file_name: str,
        old_path: Union[str, bytes, os.PathLike],
        new_path: Union[str, bytes, os.PathLike],
        new_filename: str,
    ):
        """
        Copy single data file from original location to new location, to organize data
        :param file_name: file to copy
        :param old_path: location from which to pull the data
        :param new_path: location to push the data to
        :param new_filename: used if the original filename is different from the original
        :return: copied file
        """

        old_loc = os.path.join(old_path, file_name)
        if new_filename:
            new_loc = os.path.join(new_path, new_filename)
        else:
            new_loc = os.path.join(new_path, file_name)

        try:
            os.makedirs(new_path)
        except FileExistsError:
            # directory already exists
            pass
        try:
            shutil.copy2(old_loc, new_loc)
        except shutil.SameFileError:
            pass
        except OSError:
            with open("issues.txt", "a") as f:
                f.write(f"\n{old_loc}\t{new_path}\t{file_name}")

    @staticmethod
    def get_md5_hash(
        filepath: Union[str, bytes, os.PathLike], chunk_size: int = 8192
    ) -> hashlib.md5:
        """
        Get the MD5 hash for a file, useful to verify file, or if you want to create a unique name for your file
        :param filepath: Path to file to build hash of
        :param chunk_size: bytes of the file to load at a time
        :return: md5 hash of the file
        """

        if not os.path.exists(filepath):
            raise FileNotFoundError(filepath)

        with open(filepath, "rb") as f:
            file_hash = hashlib.md5()
            chunk = f.read(chunk_size)
            while chunk:
                file_hash.update(chunk)
                chunk = f.read(chunk_size)
        return file_hash

    # ToDo: use as property when file can not be read due to encoding error
    @staticmethod
    def predict_file_encoding(
        filepath: Union[str, bytes, os.PathLike], sample_size: int = 200
    ) -> Dict:
        """
        Uses Chardet module to identify the encoding of a file.
            Will read up to the sample_size lines in the file to make the prediction
        :param filepath: Path to file
        :param sample_size: Number of lines of a file to read.
            Prevents needing to load full file into memory when working with large files.
        :return:dictionary with the predicted encoding and confidence in prediction
            {'encoding': 'utf-8', 'confidence': 0.99, 'language': ''}
        """
        assert os.path.exists(filepath)
        test_str = b""
        count = 0
        with open(filepath, "rb") as x:
            line = x.readline()
            while line and count < sample_size:
                test_str += line
                count = count + 1
                line = x.readline()

        return chardet.detect(test_str)

    # ToDo: use as a property when files will not open
    @classmethod
    def rm_bom(cls, in_file: str):
        """
        removes the BOM starting char
        :param in_file: File to be changed
        :return: None
        """
        buffer_size = 4096
        bom_length = len(codecs.BOM_UTF8)
        with open(in_file, "rb") as f:
            chunk = f.read(buffer_size)
            if chunk.startswith(codecs.BOM_UTF8):
                i = 0
                chunk = chunk[bom_length:]
                while chunk:
                    f.seek(i)
                    f.write(chunk)
                    i += len(chunk)
                    f.seek(bom_length, os.SEEK_CUR)
                    chunk = f.read(buffer_size)
                f.seek(-bom_length, os.SEEK_CUR)
                f.truncate()
        return

    @staticmethod
    def show(ingested: Union[list, dict], size: int = 2):
        """
        Used to show a few lines as example output
            Assumes normal dictionaries
        :param ingested: a list or dict to be previewed
        :param size: how many lines to view
        :return: print to display a few lines
        """

        if type(ingested) is dict:
            for c, (k, v) in enumerate(ingested.items(), 0):
                try:
                    # will show in Jupyter
                    display(k, v, "--------")
                except (KeyError, NameError):
                    print(k, v, "--------")
                if c == size:
                    break
        elif type(ingested) is defaultdict:
            for c, (k, v) in enumerate(ingested.items(), 0):
                try:
                    # Will show in Jupyter
                    display(k, dict(v), "--------")
                except (KeyError, NameError):
                    print(k, dict(v))
                if c == size:
                    break
        elif type(ingested) is list:
            for c, (i) in enumerate(ingested, 0):
                try:
                    # Will show in Jupyter
                    display(i, "--------")
                except KeyError:
                    print(i, "--------")
                if c == size:
                    break

    # noinspection PyUnboundLocalVariable
    @staticmethod
    def now(long_time: bool = False, debug_mode: bool = False) -> str:
        """
        Function to add the date in yyyy_mm_dd (hh_mm_ss)
        :param long_time: flag to show hour, minute and second info, off by default
        :param debug_mode:flag to output static date for testing purposes
        :return string_time: date as string
        """
        if type(long_time) is list:
            debug_mode = long_time[1]
            long_time = long_time[0]
        if debug_mode:
            return "20XX_02_30"
        current_time = dt.datetime.now() + dt.timedelta(seconds=0)
        if not long_time:
            string_time = current_time.strftime("%Y_%m_%d.%H_%M_%S")
        elif long_time is True:
            string_time = current_time.strftime("%Y_%m_%d")
        return str(string_time)

    @staticmethod
    def check_file(filepath: Union[str, bytes, os.PathLike]) -> str:
        """
        Check file to see if is greater than 0 bytes
        :return: 'fine' or 'problem' depending on if it is a non-empty file
        """
        file_bytes = os.stat(filepath).st_size
        if file_bytes == 0:
            return "problem"
        return "fine"

    @classmethod
    def check_for_empty(
        cls, filepath: str, bad_files: dict = None
    ) -> Union[Dict, None]:
        """
        Function to walk through multiple files and check if the data is 0 bytes or not
        :param filepath: directory to review
        :param bad_files: a dict of files which are empty
        :return: returns a dict of files which are empty or None, if applicable
        """
        bad_files = bad_files if bad_files else {}
        bad_babies = 0
        for root, dirs, files in os.walk(filepath):
            for ff in files:
                temp_file = os.path.join(root, ff)
                if cls.check_file(temp_file) == "problem":
                    bad_files[ff] = root
                    bad_babies += 1
        print(f"{bad_babies} bad files found")
        if len(bad_files) > 0:
            return bad_files
        else:
            return None

    @staticmethod
    def check_if_dict(singlet: str) -> Union[Dict, str]:
        """
        Check if the input text a dict in the form of a string escaped version of json, and converts to dict
        :param singlet: line of string, that may or may not be a json/dict item
        :return: converted string or just string if not a json/dict
        """
        try:
            out = json.loads(singlet)
        except (TypeError, json.JSONDecodeError):
            out = singlet

        return out

    @classmethod
    def open_json(
        cls, filepath: Union[str, bytes, os.PathLike], sample_output: bool = False
    ) -> Union[list, dict]:
        """
        Function to open JSON file and return array of objs
        :param sample_output: print out a sample of the file
        :param filepath: JSON file to read into memory
        :return out_file: Dict object or array of dicts
        """
        with open(filepath, "r", encoding="utf-8") as f:
            out_file = json.load(f)
            if type(out_file) == list:
                for i in out_file:
                    try:
                        for k, v in i.items():
                            try:
                                if type(v) == str:
                                    i[k] = json.loads(v)
                            except (TypeError, json.JSONDecodeError):
                                continue
                    except AttributeError as ex:
                        print(ex, type(out_file))
                        continue
            elif type(out_file) == "dict":
                for k, v in out_file.items():
                    try:
                        if type(v) == str:
                            out_file[k] = json.loads(v)
                    except (TypeError, json.JSONDecodeError):
                        continue
            if sample_output:
                try:
                    print(out_file[0])
                except KeyError:
                    for k, j in out_file.items():
                        print(k, j)
                        break
            return out_file

    @classmethod
    def open_tsv(
        cls,
        infile: Union[str, bytes, os.PathLike],
        headers: list = None,
        delimiter: str = "\t",
    ) -> Union[Dict, List]:
        """
        Open TSV file and return it as a dict or an array
        :param infile:full filepath name
        :param headers: list of headers denotes, otherwise the first row will be considered the headers
        :param delimiter: how to denote seperation between unique units, uses `\t` as default
        :return:
        """
        with open(infile, "r", encoding="utf-8") as in_filename:
            # finds the titles per column
            if headers is None:
                headers = in_filename.readline().strip().split(delimiter)

            print(f"Headers used: {headers}")
            out_array = []

            row_count = sum(1 for row in in_filename)
            in_filename.seek(0)

            dialect = csv.Sniffer().sniff(in_filename.read(1024))
            in_filename.seek(0)
            reader = csv.DictReader(
                in_filename,
                headers,
                delimiter=delimiter,
                dialect=dialect,
                quotechar='"',
                quoting=csv.QUOTE_MINIMAL,
            )
            for c, row in enumerate(reader):
                if c == 0:
                    continue
                x = {}
                for k, v in row.items():
                    x[k] = cls.check_if_dict(v)
                out_array.append(x)
            return out_array

    @classmethod
    def open_csv(
        cls,
        infile: Union[str, bytes, os.PathLike],
        headers: list = None,
    ) -> Union[Dict, List]:
        """
        Open CSV file and return it as a dict or an array
        :param infile: full filepath name
        :param headers: list of headers denotes, otherwise the first row will be considered the headers
        :return:
        """
        out_array = cls.open_tsv(infile=infile, headers=headers, delimiter=",")
        return out_array

    @classmethod
    def open_xlsx(cls, infile: Union[str, bytes, os.PathLike]) -> Union[List, Dict]:
        """ "
        Open an xlsx file and return it as a dict or an array, removing any formatting information
        :parm infile: filename of file to open
        :return: a list of dict of the information stored in the xlsx file
        """
        with open(infile, "rb") as in_filename:
            list_dict = {}
            book = openpyxl.load_workbook(in_filename, read_only=True)
            for c, name in enumerate(book.sheetnames, 0):
                list_dict[name] = []
                sheet = book.worksheets[c]
                headers = []
                for row in sheet.rows:
                    out = {}
                    for column, cell in enumerate(row, 0):
                        # Add emtpy cells to the appropriate location
                        if type(cell) == openpyxl.cell.read_only.EmptyCell:
                            out[headers[column]] = None
                            continue
                        if cell.row == 1:
                            headers.append(cell.value)
                            continue
                        else:
                            out[headers[column]] = cell.value
                    if out == {}:
                        continue
                    list_dict[name].append(out)
            if (len(list_dict.keys()) == 1) and (list_dict["Sheet"] is not None):
                return list_dict["Sheet"]
            return list_dict

    @classmethod
    def write_to_json(
        cls,
        in_array: Union[list, dict],
        lang: str = None,
        descriptor: str = None,
        location: str = None,
        long_date: bool = False,
        sample_output: bool = True,
        overwrite_name: str = None,
    ) -> str:
        """
        Function to write JSON objects to file easily
        :param in_array: JSON object to be writing to file
        :param lang: lang the data is in like en_us
        :param descriptor: indication of what the data is used for
        :param location: Location to write the file to
        :param long_date: True denotes date and time, while short denotes only date
        :param sample_output: Prints sample output to console
        :param overwrite_name: if used, will be the output name used in place of the constructed name
        :return: filename of written file
        """
        # if location not noted, save in current filepath
        if location is not None:
            original_loc = os.getcwd()
            try:
                os.chdir(location)
            except FileNotFoundError:
                os.mkdir(location)
                os.chdir(location)
        else:
            original_loc = location
        if overwrite_name:
            output_name = overwrite_name
        else:
            file_length = len(in_array)

            output_name = cls.naming(
                lang=lang,
                length=file_length,
                descriptor=descriptor,
                long_date=long_date,
                ext="json",
            )
        with open(output_name, "w", encoding="utf-8") as out_file:
            if sample_output:
                print(
                    f'Writing {file_length} items to File as "{output_name}" in \n\t"{os.getcwd()}"'
                )
            json.dump(in_array, out_file, ensure_ascii=False, indent=4)
        if location is not None:
            os.chdir(original_loc)
        return output_name

    @classmethod
    def write_to_tsv(
        cls,
        filepath: Union[str, bytes, os.PathLike],
        in_array: list,
        headers: list = None,
        delimiter: str = "\t",
    ):
        """
        Write a file to tsv with dictionaries as a row and keys asa columns
        :param filepath: Full filepath name
        :param in_array: data to be written to tsv
        :param headers: list of headers denotes, otherwise the first row will be considered the headers
        :param delimiter: denotes the seperator value
        :return:
        """
        with open(filepath, "w", newline="") as infile:
            if headers is None:
                headers = [k for k, v in in_array[0].items()]
            writing = csv.writer(
                infile, delimiter=delimiter, dialect="unix", quoting=csv.QUOTE_MINIMAL
            )
            writing.writerow(headers)
            for i in in_array:
                row = [v for k, v in i.items()]
                writing.writerow(row)
        return

    @classmethod
    def write_to_xlsx(cls, filename: Union[str, bytes, os.PathLike], in_array: list):
        """
        Created a xlsx file with dictionaries as a row and keys as columns, single sheet
        :param filename: excel filename to be use
        :param in_array: array of dicts with each obj as row
        :return: xls file with data written to it
        """
        if not filename.endswith("xlsx"):
            print("filename does not have extension xls\n...adding it")
            filename = filename + ".xlsx"
        titles = dict()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for c, i in enumerate(list(in_array[0].keys()), 1):
            titles[i] = c
            sheet.cell(row=1, column=c, value=i)

        for cc, ii in enumerate(in_array, 2):
            for k, v in ii.items():
                column = titles[k]
                sheet.cell(row=cc, column=column, value=v)

        print(f"saving to {filename}")
        workbook.save(filename)
        return filename

    @classmethod
    def tsv_to_json(
        cls,
        infile: Union[str, bytes, os.PathLike],
        headers: list = None,
        delimiter: str = "\t",
        sample_output: bool = None,
    ):
        """
        Convert a TSV into a dict or json file
        :param delimiter: denotes the seperator value
        :param infile: Full filepath name
        :param headers: list of headers denotes, otherwise the first row will be considered the headers
        :param sample_output: print out a sample of the file being converted
        :return: print out of the new filepath
        """
        _types = {"\t": "TSV", ",": "CSV"}
        out_array = cls.open_tsv(infile=infile, headers=headers, delimiter=delimiter)
        row_count = len(out_array)
        cls.write_to_json(in_array=out_array, overwrite_name=f"{infile[:-3]}json")
        if sample_output:
            print(out_array[0])
        try:
            return f"{_types[delimiter]} file written as json, with {row_count} lines as {infile[:-3]}json"
        except KeyError:
            return (
                f"TSV file written as json, with {row_count} lines as {infile[:-3]}json"
            )

    @classmethod
    def csv_to_json(
        cls,
        infile: Union[str, bytes, os.PathLike],
        headers: list = None,
        sample_output: bool = None,
    ) -> Union[str, bytes, os.PathLike]:
        """
        CSV to json conversion
        :param infile: Full filepath name
        :param headers: list of headers denotes, otherwise the first row will be considered the headers
        :param sample_output: print out a sample of the file being converted
        :return: print out of the new filepath
        """
        return cls.tsv_to_json(
            infile=infile, headers=headers, sample_output=sample_output, delimiter=","
        )

    @classmethod
    def xlsx_to_json(cls, infile: Union[str, bytes, os.PathLike]) -> dict[Any, list]:
        """
        Converts an excel-like file (xlsx) to a json object with each sheet being the key with one dict per Row
            First row is assumed to be headers
                'Test_Sheet1': [
                    {'Column 1': 'Row 1', 'Column 2': 'Row1A'},
                    {'Column 1': 'Row 2', 'Column 2': 'Row2A'}
                    ]
        :param infile: excel book to open and convert
        :return: Write to file with the same name in the same location
        """
        list_dict = cls.open_xlsx(infile)
        outpath_name = f"{infile[:-4]}json"
        cls.write_to_json(list_dict, overwrite_name=outpath_name)
        return

    @classmethod
    def pull_data(cls, folder_location: str) -> Dict:
        """
        Stroll through a folder to pull out data
        :param folder_location: Folder to pull data from
        :return: data stored as a dict with filename as key, and contents as value
        """
        all_data = {}
        for root, dirs, files in os.walk(folder_location):
            for c, i in enumerate(files):
                if i == "":
                    continue
                    # create dict for item if it doesn't already exist
                filename, ext = [x.lower() for x in os.path.splitext(i)]
                if ext == ".xlsx":
                    temp = cls.open_xlsx(f"{root}/{i}")
                elif ext == ".json":
                    temp = cls.open_json(f"{root}/{i}")
                elif ext == ".tsv":
                    cls.open_tsv(f"{root}/{i}")
                elif ext == ".csv":
                    cls.open_csv(infile=f"{root}/{i}")
                else:
                    print(f"Unable to open file {i}")
                    temp = []
            all_data[i] = temp

        return all_data
